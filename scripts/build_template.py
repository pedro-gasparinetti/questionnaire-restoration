"""Build the export template XLSX with a Data sheet and a Metadata sheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# -----------------------------------------------------------------------------
# Sheet 1: Data
# -----------------------------------------------------------------------------
ws = wb.active
ws.title = "Data"

cols = []
# 1. Identification
cols += ["Respondent", "User", "Date", "GPS", "Ecosystem", "Country", "City"]
# 2. Method
cols += ["Method", "Method_ID"]
# 3. Implementation
cols += ["Impl_USD", "Impl_Labor_%", "Impl_Mater_%", "Impl_Mach_%"]
# G. Maint segments
activities = ["RegInd", "MaintNTFP", "Harvest", "TechAssist", "Monitoring"]
for act in activities:
    for n in range(1, 11):
        cols += [
            f"Maint_{act}_Seg{n}_From_yr",
            f"Maint_{act}_Seg{n}_To_yr",
            f"Maint_{act}_Seg{n}_Annual_USD",
        ]
# F. Maint distribution
cols += ["Maint_Labor_%", "Maint_Mater_%", "Maint_Mach_%"]
# 6. NTFP species & price
cols += ["NTFP_Species", "NTFP_Price_USD_kg"]
# I. ProdSeg
for n in range(1, 6):
    cols += [f"ProdSeg{n}_From_yr", f"ProdSeg{n}_To_yr", f"ProdSeg{n}_kg_ha_yr"]
# J. RevSeg
for n in range(1, 6):
    cols += [f"RevSeg{n}_From_yr", f"RevSeg{n}_To_yr", f"RevSeg{n}_Annual_USD"]
# B. Context constraints
cols += [
    "Fire_UnitCost_USD_km", "Fire_UnitCost_USD_ha", "Fire_Occur",
    "Fire_FirebreakArea_ha", "Fire_Labor_%", "Fire_Mater_%", "Fire_Mach_%",
]
cols += [
    "Fence_UnitCost_USD_km", "Fence_UnitCost_USD_ha", "Fence_Area_ha",
    "Fence_Labor_%", "Fence_Mater_%", "Fence_Mach_%",
]
cols += ["Weed_UnitCost", "Weed_Occur", "Weed_Labor_%", "Weed_Mater_%", "Weed_Mach_%"]
cols += ["Pest_UnitCost", "Pest_Occur", "Pest_Labor_%", "Pest_Mater_%", "Pest_Mach_%"]
# C. Labor breakdown
cols += ["Impl_Hired_%", "Impl_Family_%", "Maint_Hired_%", "Maint_Family_%"]
cols += ["HiredLaborCost_USD_day", "MachineryUnitCost_USD_hr", "LandLease_USD_ha_yr"]
cols += ["Gender_Male_%", "Gender_Female_%", "Gender_Other_%"]

header_font = Font(bold=True, color="FFFFFF", name="Arial")
header_fill = PatternFill("solid", start_color="2A4B46")
center = Alignment(horizontal="center", vertical="center")

for i, name in enumerate(cols, 1):
    cell = ws.cell(row=1, column=i, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    ws.column_dimensions[cell.column_letter].width = 22

ws.freeze_panes = "A2"

methods = [
    ("ANR/50% Enrichment", "anr_30"),
    ("ANR/50% Enrichment (with NTFP)", "anr_30_ntfp"),
    ("Seed Dispersal", "seed_dispersal"),
    ("Seed Dispersal (with NTFP)", "seed_dispersal_ntfp"),
    ("Full Seedling Plantation", "seedling_planting"),
    ("Full Seedling Plantation NTFP (agroforestry)", "seedling_planting_ntfp"),
]
for r, (label, mid) in enumerate(methods, start=2):
    ws.cell(row=r, column=cols.index("Method") + 1, value=label)
    ws.cell(row=r, column=cols.index("Method_ID") + 1, value=mid)

# -----------------------------------------------------------------------------
# Sheet 2: Metadata
# -----------------------------------------------------------------------------
mws = wb.create_sheet("Metadata")

mws.cell(row=1, column=1, value="Field").font = header_font
mws.cell(row=1, column=2, value="Unit").font = header_font
mws.cell(row=1, column=3, value="Section").font = header_font
mws.cell(row=1, column=4, value="Description").font = header_font
for c in range(1, 5):
    mws.cell(row=1, column=c).fill = header_fill
    mws.cell(row=1, column=c).alignment = center

meta = [
    # 1. Identification
    ("Respondent", "text", "1. Identification", "Name of the field expert / consultant who provided the data"),
    ("User", "text", "1. Identification", "Name of the person filling out the questionnaire"),
    ("Date", "YYYY-MM-DD", "1. Identification", "Date when the data was collected"),
    ("GPS", 'text "lat, lon"', "1. Identification", "GPS coordinates of the project site"),
    ("Ecosystem", "text", "1. Identification", "Ecosystem type (e.g., Tropical Forest, Cerrado, Mangrove)"),
    ("Country", "text", "1. Identification", "Country where the project is located"),
    ("City", "text", "1. Identification", "City where the project is located"),
    # 2. Method
    ("Method", "text", "2. Method", "Display label of the restoration method for this row"),
    ("Method_ID", "text", "2. Method", "Internal ID of the method (anr_30, anr_30_ntfp, seed_dispersal, seed_dispersal_ntfp, seedling_planting, seedling_planting_ntfp)"),
    # 3. Implementation
    ("Impl_USD", "US$/ha", "3. Implementation", "Basic implementation cost (Year 1) for this method"),
    ("Impl_Labor_%", "%", "3. Implementation", "Share of implementation cost attributable to labor (must sum to 100% with Mater/Mach)"),
    ("Impl_Mater_%", "%", "3. Implementation", "Share of implementation cost attributable to materials"),
    ("Impl_Mach_%", "%", "3. Implementation", "Share of implementation cost attributable to machinery/services"),
    # G. Maint segments — one row per activity
    (
        "Maint_RegInd_Seg<N>_<Field>",
        "year (From/To) or US$/ha/yr (Annual_USD)",
        "4. Maintenance segments",
        "Segments for 'Maintenance of regenerating individuals' (pruning, thinning, support staking). <N> in 1..10. <Field> in {From_yr (start year), To_yr (end year, >= From_yr), Annual_USD (annual cost in US$/ha)}. Empty when the user has not used that segment slot.",
    ),
    (
        "Maint_MaintNTFP_Seg<N>_<Field>",
        "year (From/To) or US$/ha/yr (Annual_USD)",
        "4. Maintenance segments",
        "Segments for 'Maintenance of NTFP species' (pruning, thinning, support staking). <N> in 1..10. <Field> in {From_yr, To_yr, Annual_USD}. Empty when the slot was not used.",
    ),
    (
        "Maint_Harvest_Seg<N>_<Field>",
        "year (From/To) or US$/ha/yr (Annual_USD)",
        "4. Maintenance segments",
        "Segments for 'Harvest' activity (e.g. NTFP collection, when treated as a maintenance cost). <N> in 1..10. <Field> in {From_yr, To_yr, Annual_USD}. Empty when the slot was not used.",
    ),
    (
        "Maint_TechAssist_Seg<N>_<Field>",
        "year (From/To) or US$/ha/yr (Annual_USD)",
        "4. Maintenance segments",
        "Segments for 'Technical assistance' (agronomist/forester visits, planning, training). <N> in 1..10. <Field> in {From_yr, To_yr, Annual_USD}. Empty when the slot was not used.",
    ),
    (
        "Maint_Monitoring_Seg<N>_<Field>",
        "year (From/To) or US$/ha/yr (Annual_USD)",
        "4. Maintenance segments",
        "Segments for 'Monitoring General Maintenance Activities' (e.g. shade management, light interventions). <N> in 1..10. <Field> in {From_yr, To_yr, Annual_USD}. Empty when the slot was not used.",
    ),
    # F. Maint distribution
    ("Maint_Labor_%", "%", "5. Maintenance distribution", "Share of maintenance cost attributable to labor (sum to 100%)"),
    ("Maint_Mater_%", "%", "5. Maintenance distribution", "Share of maintenance cost attributable to materials"),
    ("Maint_Mach_%", "%", "5. Maintenance distribution", "Share of maintenance cost attributable to machinery/services"),
    # 6. NTFP species & price
    ("NTFP_Species", "text", "6. NTFP species & price", "Selected Non-Timber Forest Product species (NTFP methods only - empty otherwise)"),
    ("NTFP_Price_USD_kg", "US$/kg", "6. NTFP species & price", "Average price of the NTFP during harvesting season (NTFP methods only)"),
    # I. ProdSeg
    (
        "ProdSeg<N>_<Field>",
        "year (From/To) or kg/ha/yr (kg_ha_yr)",
        "7. NTFP Productivity segments",
        "NTFP productivity segments. <N> in 1..5. <Field> in {From_yr, To_yr, kg_ha_yr (annual productivity)}. Filled only when the user picks 'Productivity data' mode for an NTFP method. Empty otherwise.",
    ),
    # J. RevSeg
    (
        "RevSeg<N>_<Field>",
        "year (From/To) or US$/ha/yr (Annual_USD)",
        "8. NTFP Revenue segments",
        "NTFP revenue segments. <N> in 1..5. <Field> in {From_yr, To_yr, Annual_USD (annual revenue)}. Filled only when the user picks 'Revenue data' mode for an NTFP method. Empty otherwise.",
    ),
    # B. Context constraints - Fire
    ("Fire_UnitCost_USD_km", "US$/km", "9. Context Constraints", "Firebreak unit cost per linear km"),
    ("Fire_UnitCost_USD_ha", "US$/ha", "9. Context Constraints", "Firebreak unit cost per hectare (alternative input - converted to per-km internally)"),
    ("Fire_Occur", "count", "9. Context Constraints", "Number of times firebreak activity occurs over the 20-year horizon"),
    ("Fire_FirebreakArea_ha", "ha", "9. Context Constraints", "Average total area that needs fire breaks (used for US$/ha to US$/km conversion)"),
    ("Fire_Labor_%", "%", "9. Context Constraints", "Share of firebreak cost attributable to labor (sum to 100%)"),
    ("Fire_Mater_%", "%", "9. Context Constraints", "Share of firebreak cost attributable to materials"),
    ("Fire_Mach_%", "%", "9. Context Constraints", "Share of firebreak cost attributable to machinery/services"),
    # B - Fence
    ("Fence_UnitCost_USD_km", "US$/km", "9. Context Constraints", "Fencing unit cost per linear km"),
    ("Fence_UnitCost_USD_ha", "US$/ha", "9. Context Constraints", "Fencing unit cost per hectare (alternative input - converted to per-km internally)"),
    ("Fence_Area_ha", "ha", "9. Context Constraints", "Average area that needs fences in one typical property"),
    ("Fence_Labor_%", "%", "9. Context Constraints", "Share of fencing cost attributable to labor"),
    ("Fence_Mater_%", "%", "9. Context Constraints", "Share of fencing cost attributable to materials"),
    ("Fence_Mach_%", "%", "9. Context Constraints", "Share of fencing cost attributable to machinery/services"),
    # B - Weed
    ("Weed_UnitCost", "US$/ha", "9. Context Constraints", "Invasive species / weed control unit cost"),
    ("Weed_Occur", "count", "9. Context Constraints", "Number of weed-control occurrences over the 20-year horizon"),
    ("Weed_Labor_%", "%", "9. Context Constraints", "Share of weed-control cost attributable to labor"),
    ("Weed_Mater_%", "%", "9. Context Constraints", "Share of weed-control cost attributable to materials"),
    ("Weed_Mach_%", "%", "9. Context Constraints", "Share of weed-control cost attributable to machinery/services"),
    # B - Pest
    ("Pest_UnitCost", "US$/ha", "9. Context Constraints", "Pest control unit cost"),
    ("Pest_Occur", "count", "9. Context Constraints", "Number of pest-control occurrences over the 20-year horizon"),
    ("Pest_Labor_%", "%", "9. Context Constraints", "Share of pest-control cost attributable to labor"),
    ("Pest_Mater_%", "%", "9. Context Constraints", "Share of pest-control cost attributable to materials"),
    ("Pest_Mach_%", "%", "9. Context Constraints", "Share of pest-control cost attributable to machinery/services"),
    # C. Labor breakdown
    ("Impl_Hired_%", "%", "10. Labor Breakdown", "Implementation labor: share of hired (paid) workers (Hired+Family = 100%)"),
    ("Impl_Family_%", "%", "10. Labor Breakdown", "Implementation labor: share of family (unpaid) labor"),
    ("Maint_Hired_%", "%", "10. Labor Breakdown", "Maintenance labor: share of hired (paid) workers"),
    ("Maint_Family_%", "%", "10. Labor Breakdown", "Maintenance labor: share of family (unpaid) labor"),
    ("HiredLaborCost_USD_day", "US$/day", "10. Labor Breakdown", "Regional reference: average daily wage for hired field workers"),
    ("MachineryUnitCost_USD_hr", "US$/hour", "10. Labor Breakdown", "Regional reference: hourly cost of machinery"),
    ("LandLease_USD_ha_yr", "US$/ha/year", "10. Labor Breakdown", "Regional reference: average annual land lease (rent) rate per hectare"),
    ("Gender_Male_%", "%", "10. Labor Breakdown", "Share of total labor hours contributed by male workers (sum to 100% with Female/Other)"),
    ("Gender_Female_%", "%", "10. Labor Breakdown", "Share of total labor hours contributed by female workers"),
    ("Gender_Other_%", "%", "10. Labor Breakdown", "Share of total labor hours contributed by non-binary / other workers"),
]

wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)

for r, (field, unit, section, desc) in enumerate(meta, start=2):
    mws.cell(row=r, column=1, value=field).font = Font(name="Consolas", size=10)
    mws.cell(row=r, column=2, value=unit).font = Font(name="Arial", size=10)
    mws.cell(row=r, column=3, value=section).font = Font(name="Arial", size=10, italic=True, color="666666")
    mws.cell(row=r, column=4, value=desc).font = Font(name="Arial", size=10)
    for c in range(1, 5):
        mws.cell(row=r, column=c).alignment = wrap
        mws.cell(row=r, column=c).border = border

mws.column_dimensions["A"].width = 38
mws.column_dimensions["B"].width = 28
mws.column_dimensions["C"].width = 30
mws.column_dimensions["D"].width = 90
mws.freeze_panes = "A2"

print(f"Data sheet: {len(cols)} columns")
print(f"Metadata sheet: {len(meta)} rows")

out = r"C:\Users\maria\OneDrive\Documentos\Trabalho\Consultorias\CSF\2025\CI - Calculadora\Questionnaires\export_template_v5.xlsx"
wb.save(out)
print(f"Saved to: {out}")
